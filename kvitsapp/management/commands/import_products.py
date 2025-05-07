# Palaiž programmu, kas importē datus no Excel faila
# docker-compose exec web python manage.py import_products

# Izdzēst datus datubāzē
# docker-compose exec web python manage.py flush

# Atjaunina datubāzi, ja ir izmaiņas modeļos
# docker-compose exec web python manage.py makemigrations kvitsapp
# docker-compose exec web python manage.py migrate

# Lai apskatītu datubāzi jaiziet caur šiem soļiem:
# docker-compose exec web bash
    # Vienreiž jāpalaiž lai būtu pieejams psql
    # apt-get update
    # apt-get install -y postgresql-client
# psql -h db -U postgres
# postgres
# SELECT * FROM kvitsapp_product;
# Lai izietu no psql
# \q

import os
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from django.core.management.base import BaseCommand
from django.conf import settings
from kvitsapp.models import Product, Category
from decimal import Decimal
from django.utils.text import slugify

# This class inherits from BaseCommand and contains the core logic for the import process.
class Command(BaseCommand):
    help = 'Import products from Excel and extract images'

    # PART 2A: _preprocess_image_mapping method
    # This helper function analyzes the Excel sheet to create a map that associates
    # rows with image locations (cell addresses). It's designed to handle images
    # within merged cells and correctly link them to the relevant product rows.
    def _preprocess_image_mapping(self, sheet, image_loader):
        row_to_image = {}
        image_cells = {}
        # First, identify all cells that contain an image
        for row in range(1, sheet.max_row + 1):
            for cell in sheet[row]:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell): # Skip merged cells initially
                    cell_address = cell.coordinate
                    if image_loader.image_in(cell_address):
                        image_cells[cell_address] = (row, cell.column)

        # Now, map rows to these image cells, considering merged cells
        for cell_address, (img_row, img_col) in image_cells.items():
            is_part_of_merge = False
            # Check if the image cell is the top-left of a merged range
            for merged_range in sheet.merged_cells.ranges:
                # Get the top-left cell of the merged range
                start_cell_in_merge = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                if start_cell_in_merge.coordinate == cell_address:
                    # If it is, associate all rows in this merged range with this image
                    for r in range(merged_range.min_row, merged_range.max_row + 1):
                        row_to_image[r] = cell_address
                    is_part_of_merge = True
                    break
            
            if not is_part_of_merge:
                # If not part of a merge, or if it's a single cell image, map the current row
                row_to_image[img_row] = cell_address
                # Heuristic: Try to associate this image with subsequent rows if they don't have their own
                # and appear to be related (e.g., have data in key columns like product code).
                # This helps if an image is meant for a few subsequent product entries.
                for r in range(img_row + 1, min(img_row + 5, sheet.max_row + 1)): # Look ahead a few rows
                    if r not in row_to_image: # If the row isn't already mapped
                        # Check if the row has product data (e.g., in first or second column)
                        if sheet.cell(row=r, column=1).value or sheet.cell(row=r, column=2).value:
                            # Ensure this row doesn't get its own image assigned later from a closer source
                            has_own_image_closer = False
                            for check_r in range(r, min(r + 3, sheet.max_row + 1)): # Check a small window
                                if check_r in row_to_image and check_r != img_row: # and check_r has an image from a different source
                                    # A more specific image might be found for 'check_r' later in image_cells,
                                    # so we need a more robust way to check if 'r' will get its own image.
                                    # This part of the logic is complex and heuristic.
                                    # For now, this simpler check might lead to over-assigning if not careful.
                                    pass # This logic can be refined
                            if not has_own_image_closer: # A simplified check for now
                                row_to_image[r] = cell_address
                        # else: # Row is likely empty or a continuation without its own image
                            # break # Stop associating if we hit an empty-looking row
        return row_to_image

    # PART 2B: _find_image_for_row method
    # This function attempts to locate the appropriate image for a specific product row
    # using the pre-processed map. If a direct match isn't found in the map,
    # it searches nearby preceding rows for an applicable image.
    def _find_image_for_row(self, sheet, image_loader, row_index, row_to_image_map):
        # Check the pre-processed map first
        if row_index in row_to_image_map:
            cell_address = row_to_image_map[row_index]
            # Determine if this cell_address corresponds to a merged cell containing this row_index
            is_merged_for_this_row = False
            for merged_range in sheet.merged_cells.ranges:
                if (merged_range.min_row <= row_index <= merged_range.max_row):
                    top_left_merged_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).coordinate
                    if top_left_merged_cell == cell_address:
                        is_merged_for_this_row = True
                        break
            return cell_address, is_merged_for_this_row

        # Fallback: if not in map, iterate backwards from current row to find the nearest preceding image.
        # This might be relevant if an image implicitly applies to multiple following rows
        # and the pre-processing didn't catch it, or for single-cell images not in merged ranges.
        for r in range(row_index, max(1, row_index - 10), -1): # Look back up to 10 rows
            for cell in sheet[r]: # Check all cells in that row
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell_address = cell.coordinate
                    if image_loader.image_in(cell_address):
                        # Found an image in a preceding row.
                        # 'is_merged' here would mean the image found is from a different row's context.
                        return cell_address, r != row_index 
        return None, False # No image found

    # PART 2C: handle method
    # This is the main entry point for the Django management command.
    # It orchestrates the entire import process.
    def handle(self, *args, **kwargs):
        # Define input Excel file path and output directory for extracted images.
        filepath = os.path.join('katalogs_2024.xlsx')
        output_folder = os.path.join(settings.BASE_DIR, 'kvitsapp', 'static', 'images')
        os.makedirs(output_folder, exist_ok=True) # Create images folder if it doesn't exist

        # Category map: Assigns categories to products based on prefixes in their order codes.
        category_map = {
            "EL": "Klasiskās",
            "E100"  : "Klasiskās",
            "EM40"  : "Klasiskās",
            "EM50"  : "Klasiskās",
            "EM200"  : "Klasiskās",
            "E138"  : "Klasiskās",
            "E300"  : "Klasiskās",
            "E200"  : "Klasiskās",
            "EV": "Vārtu",
            "EM": "Metināmās",
            "ET": "T veida",
            "EK": "Klavieru",
            "EA100": "Atsper",
            "EA": "Antīkās",
            "DEKORS": "Antīkās",

            "AZ": "Aizbīdņi",
            "AA": "Aizbīdņi",
            "AZV": "Vārtiem",
            "AZ50": "Vārtiem",
            "AZ750": "Vārtiem",
            "AZ30": "Vārtiem",
            "S204": "Vārtiem",
            "AZ4": "Vārtiem",
            "AZ38": "Vārtiem",
            "KR": "Kronšteini",
            "AI": "Iekaļamie aizbīdņi",
            "KR100": "Krampji",
            "KR127": "Krampji",
            "KR250": "Krampji",

            "R21": "Garajām uzlikām",
            "R51": "Garajām uzlikām",
            "R31": "Garajām uzlikām",
            "R06": "Dalītajām uzlikām",
            "R26": "Dalītajām uzlikām",
            "R182": "Dalītajām uzlikām",
            "R199": "Dalītajām uzlikām",
            "R52": "Dalītajām uzlikām",
            "R41": "Dalītajām uzlikām",
            "R103S": "Dalītajām uzlikām",
            "R106S": "Dalītajām uzlikām",
            "R640": "Skandināvu rokturi",
            "R75": "Skavveida",
            "R100": "Skavveida",
            "R125": "Skavveida",
            "R150": "Skavveida",
            "R160": "Skavveida",
            "R162": "Skavveida",	
            "R175": "Skavveida",
            "R200": "Skavveida",
            "R212": "Skavveida",
            "R225": "Skavveida",
            "R230": "Skavveida",
            "R100MM": "Stieņi",
            "R150MM": "Stieņi",
            "RC": "Centra",
            "RL": "Lūkas",
            "RKOKA": "Koka",

            "S0": "Pretplāksnes",
            "S2": "Vācu standarta",
            "S.B": "Uzliekamās",
            "S45": "Starpistabu",
            "S155": "Rulīšu_mehanismi",
            "S55": "WC slēdzenes",
            "ASSA": "Skandināvu standarta",
            "S114": "Skandināvu standarta",
            "S40": "Universālās",
            "S85": "Eiro standarta",
            "S2018": "Profilcilindram",

            "C3": "Parastie",
            "CA": "Parastie",
            "C3": "Parastie",
            "C4": "Parastie",
            "CM": "Multi cilindri",
            "R0": "Uzlikas",
            "R113": "Uzlikas",
            "R115": "Uzlikas",
            "R02P": "Uzlikas",
            "R01PZSS": "Uzlikas",
            "R111": "WC",
            "R02": "WC",
            "R01": "WC",

            "SM": "Mēbeļu slēdzenes",
            "M5": "Magnēti",
            "L00": "Lodītes",

            "MB": "Margu balsti",
            "K": "Konsoles",

            "A20": "Actiņas",   
            "PD": "Piekaramās atslēgas",
            "N": "Numuriņi",
            "AT": "Atsperes",
            "KL": "Durvju aizvērēji",
            "A8": "Durvju aizvērēji",
            "PK": "Pakaramie",
            "KR4": "Pakaramie",
            "KR5": "Pakaramie",
            "A0": "Atdures",

        }

        try:
            # Load the Excel workbook and initialize the image loader.
            wb = openpyxl.load_workbook(filepath, data_only=True) # data_only=True to get values, not formulas
            sheet = wb.active
            image_loader = SheetImageLoader(sheet)
            
            self.stdout.write(self.style.SUCCESS('Pre-processing sheet to map rows to images...'))
            row_to_image_map = self._preprocess_image_mapping(sheet, image_loader)
            self.stdout.write(self.style.SUCCESS(f'Pre-processing complete. Found {len(row_to_image_map)} potential row-to-image associations.'))

            # processed_images = {} # To track if an image from a specific cell has been saved
            cell_to_filename = {} # Maps a cell address (image source) to its saved filename to avoid duplicates

            # Iterate through each row of the spreadsheet (skipping the header).
            for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2): # min_row=2 to skip header
                try:
                    # Extract product details.
                    ean13 = row[0].value
                    pasutijuma_kods = row[1].value
                    if not pasutijuma_kods: # Skip row if no order code (essential identifier)
                        self.stdout.write(self.style.WARNING(f'Skipping row {row_index} due to missing Pasūtijuma Kods.'))
                        continue

                    apraksts = row[3].value
                    cena_raw_value = row[6].value

                    # Process price
                    if cena_raw_value is not None:
                        try:
                            cena = Decimal(str(cena_raw_value).replace(',', '.'))
                        except (ValueError, TypeError):
                            self.stdout.write(self.style.WARNING(f'Skipping row {row_index} (Kods: {pasutijuma_kods}) due to invalid price: {cena_raw_value}'))
                            continue
                    else:
                        self.stdout.write(self.style.WARNING(f'Skipping row {row_index} (Kods: {pasutijuma_kods}) due to missing price.'))
                        continue

                    # Determine product category.
                    kod = str(pasutijuma_kods)
                    # Try matching with progressively shorter prefixes to find the most specific category.
                    prefixes = [kod[:i] for i in range(min(7, len(kod)), 0, -1)] 

                    category_name = "Citi" # Default category
                    for p in prefixes:
                        if p in category_map:
                            category_name = category_map[p]
                            break
                    
                    category_slug = slugify(category_name)
                    category, _ = Category.objects.get_or_create(name=category_name, slug=category_slug)

                    # Find and process the associated image.
                    # cell_address will be the coordinate of the cell containing the image (e.g., 'A1')
                    # is_merged indicates if this image was found in a merged cell context relative to the current row
                    cell_address, is_merged_context = self._find_image_for_row(sheet, image_loader, row_index, row_to_image_map)
                    attels_filename = "images/no-image.png" # Default placeholder

                    if cell_address:
                        # If this image (from this specific cell_address) has already been processed and saved, reuse its filename.
                        if cell_address in cell_to_filename:
                            attels_filename = cell_to_filename[cell_address]
                            # self.stdout.write(f"Row {row_index}: Reusing image from {cell_address} for {pasutijuma_kods}")
                        else:
                            # New image to process from this cell_address.
                            try:
                                image = image_loader.get(cell_address)
                                # Sanitize pasutijuma_kods for filename
                                safe_kods = "".join(c if c.isalnum() else "_" for c in str(pasutijuma_kods))
                                filename = f"product_{safe_kods}.png"
                                image_path = os.path.join(output_folder, filename)
                                image.save(image_path)
                                attels_filename = f"images/{filename}" # Path relative to static folder
                                cell_to_filename[cell_address] = attels_filename # Store for reuse
                                # self.stdout.write(f"Row {row_index}: Saved new image from {cell_address} as {filename} for {pasutijuma_kods}")
                            except Exception as img_e:
                                self.stdout.write(self.style.ERROR(f'Error saving image from cell {cell_address} for row {row_index} (Kods: {pasutijuma_kods}): {img_e}'))
                                attels_filename = "images/no-image.png"
                    # else:
                        # self.stdout.write(f"Row {row_index}: No image found for {pasutijuma_kods}")


                    # Create or update a Product record in the database.
                    Product.objects.update_or_create(
                        pasutijuma_kods=pasutijuma_kods,
                        defaults={
                            'ean13': ean13 or '',
                            'attels': attels_filename,
                            'apraksts': apraksts or '',
                            'cena': cena,
                            'category': category.name # Storing category name, assuming model handles relation
                        }
                    )

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'Error processing row {row_index} (Kods: {pasutijuma_kods if "pasutijuma_kods" in locals() else "N/A"}): {e}'))

            self.stdout.write(self.style.SUCCESS(f'Products imported successfully! Processed {len(cell_to_filename)} unique images.'))

        # PART 3: ERROR HANDLING
        # Includes error handling for file not found and other general exceptions,
        # outputting status messages to the console.
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'Error: The file "{filepath}" was not found.'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'An unexpected error occurred: {e}'))